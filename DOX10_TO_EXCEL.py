#USE FILECLEANER FIRST

refined_list = []

f = open("IO_TXT.txt", "r",  encoding='unicode_escape')
linenum = 0
for line in f:
    first = line.strip(" ")
    second = first.split(" ")
    temporary_list = []

    for i in second:
        if i != '':
            temporary_list.append(i)

    refined_list.append(temporary_list)
    linenum += 1
print(str(linenum - 1) + "values was added")


from openpyxl import Workbook

wb = Workbook()

dest_filename = 'IO_EXL.xlsx'
ws1 = wb.active
ws1.title = "IO DOX10 Ørin"
ws1['A1'] = "Variable"
ws1['B1'] = "IO"
ws1['C1'] = "Identifier"
ws1['D1'] = "Comment"


i_num = 2
for i in refined_list:
    try:
        ws1['A' + str(i_num)] = str(i[0])
    except:
        ws1['A' + str(i_num)] = "NONE"
    try:
        #print(i[1])
        ws1['B' + str(i_num)] = str(i[1])
    except:
        ws1['B' + str(i_num)] = "NONE"
    try:
        ws1['C' + str(i_num)] = str(i[2])
    except:
        ws1['C' + str(i_num)] = " "

    comment_string = ""
    comment = ""
    for c in i[3:]:
        comment_string += str(c) + " "
    for l in comment_string:
        if l != "\n":
            if l == '\x9b':
                comment += "ø"
            if l == '\x8f':
                comment += "Å"
            if l == '\x86':
                comment += "å"
            if l == '\x9d':
                comment += "Ø"
            else:
                comment += str(l)
        else:
            pass
    try:
        ws1['D' + str(i_num)] = comment
    except:
        ws1['D' + str(i_num)] = " "
    print(comment)
    i_num += 1

wb.save(filename = dest_filename)



#ws2 = wb.create_sheet(title="Pi")

#ws2['F5'] = 3.14

#ws3 = wb.create_sheet(title="Data")

#print(ws3['AA10'].value)

